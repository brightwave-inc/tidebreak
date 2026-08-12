#!/usr/bin/env python3
"""Inventory an XLSX workbook, sample rows, and thumbnail its first sheets."""

from __future__ import annotations

import argparse
from pathlib import Path

from _tidebreak_preview import (
    HelperError,
    ensure_preview_dir,
    existing_file,
    overview_grid,
    remove_matching,
    require_pillow,
    run_cli,
    slug,
)


def parser() -> argparse.ArgumentParser:
    cli = argparse.ArgumentParser(
        description="Print XLSX sheet ranges/sample rows and write thumbnails into preview/.",
    )
    cli.add_argument("workbook", help="XLSX file inside the exec workspace")
    cli.add_argument(
        "--sample-rows",
        type=int,
        default=5,
        help="non-empty rows printed per sheet, from 1 through 20 (default: 5)",
    )
    cli.add_argument(
        "--max-sheets",
        type=int,
        default=2,
        help="sheet thumbnails to emit, from 1 through 8 (default: 2)",
    )
    cli.add_argument(
        "--preview-dir",
        default="preview",
        help="preview output directory (default: preview)",
    )
    return cli


def _openpyxl():
    try:
        import openpyxl
    except ImportError as error:
        raise HelperError(
            "XLSX analysis requires the Python 'openpyxl' package"
        ) from error
    return openpyxl


def _display(value) -> str:
    if value is None:
        return ""
    rendered = str(value).replace("\r", " ").replace("\n", " ")
    return rendered[:80] + ("…" if len(rendered) > 80 else "")


def _sample_rows(sheet, count: int, columns: int) -> list[list[str]]:
    samples: list[list[str]] = []
    for row in sheet.iter_rows(
        min_row=1,
        max_row=min(sheet.max_row, 500),
        min_col=1,
        max_col=min(max(columns, 1), 12),
        values_only=True,
    ):
        values = [_display(value) for value in row]
        if any(values):
            samples.append(values)
            if len(samples) == count:
                break
    return samples


def _sheet_thumbnail(name: str, samples: list[list[str]], destination: Path) -> None:
    Image, ImageDraw, _ImageOps = require_pillow()
    columns = max(1, min(max((len(row) for row in samples), default=1), 8))
    rows = max(1, min(len(samples), 10))
    cell_width = 150
    cell_height = 34
    title_height = 42
    image = Image.new(
        "RGB",
        (columns * cell_width, title_height + rows * cell_height),
        "white",
    )
    draw = ImageDraw.Draw(image)
    draw.rectangle((0, 0, image.width - 1, title_height - 1), fill="#e9eef6")
    draw.text((10, 13), name[:80], fill="#172033")
    for row_index in range(rows):
        values = samples[row_index] if row_index < len(samples) else []
        for column_index in range(columns):
            x = column_index * cell_width
            y = title_height + row_index * cell_height
            fill = "#f7f9fc" if row_index % 2 else "white"
            draw.rectangle(
                (x, y, x + cell_width - 1, y + cell_height - 1),
                fill=fill,
                outline="#c8ced8",
            )
            value = values[column_index] if column_index < len(values) else ""
            draw.text((x + 6, y + 10), value[:22], fill="#172033")
    image.save(destination, format="PNG", optimize=True)


def main() -> int:
    args = parser().parse_args()
    if not 1 <= args.sample_rows <= 20:
        raise HelperError("sample-rows must be between 1 and 20")
    if not 1 <= args.max_sheets <= 8:
        raise HelperError("max-sheets must be between 1 and 8")
    source = existing_file(args.workbook, [".xlsx", ".xlsm"])
    preview = ensure_preview_dir(args.preview_dir)
    openpyxl = _openpyxl()
    require_pillow()

    workbook = openpyxl.load_workbook(
        source,
        read_only=True,
        data_only=True,
    )
    remove_matching(preview, ["overview-grid.png", "sheet-*.png"])
    print(f"Workbook: {source.name}")
    print(f"Sheets: {len(workbook.sheetnames)}")
    thumbnails: list[tuple[Path, str]] = []
    for index, name in enumerate(workbook.sheetnames, start=1):
        sheet = workbook[name]
        used_range = sheet.calculate_dimension(force=True)
        rows = max(sheet.max_row or 0, 0)
        columns = max(sheet.max_column or 0, 0)
        samples = _sample_rows(sheet, args.sample_rows, columns)
        print(f"- {name}: used={used_range}, rows={rows}, columns={columns}")
        if samples:
            for sample_index, values in enumerate(samples, start=1):
                print(f"  sample {sample_index}: " + " | ".join(values))
        else:
            print("  sample: <empty>")
        if index <= args.max_sheets:
            output = preview / f"sheet-{index:03d}-{slug(name)}.png"
            _sheet_thumbnail(name, samples, output)
            thumbnails.append((output, name))

    if thumbnails:
        overview = preview / "overview-grid.png"
        overview_grid(thumbnails, overview)
        names = ", ".join(
            path.name for path, _label in [(overview, "Overview"), *thumbnails]
        )
        print(f"Preview images: {names}")
    if len(workbook.sheetnames) > len(thumbnails):
        print(
            f"Thumbnail limit skipped {len(workbook.sheetnames) - len(thumbnails)} "
            "sheet(s); rerun with --max-sheets after narrowing the workbook."
        )
    workbook.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(run_cli(main))
